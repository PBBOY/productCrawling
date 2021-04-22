import random
import logging
import time
import math
import csv
import logging

from openpyxl import Workbook

from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel
from starlette.responses import JSONResponse


def create_aliased_response(model: BaseModel) -> JSONResponse:
    return JSONResponse(content=jsonable_encoder(model, by_alias=True))


class Singleton(type):
    _instances = {}

    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super().__call__(*args, **kwargs)
        else:
            pass
            # 매번 __init__ 호출하고 싶으면 아래 주석 제거
            # cls._instances[cls].__init__(*args, **kwargs)
        return cls._instances[cls]


class Utils:
    @classmethod
    def take_a_sleep(cls, s: int, e: int):
        """S ~ E 사이의 랜던값으로 Sleep"""
        random_count = random.uniform(s, e)
        # logging.info('take a sleep: '+str(random_count))
        time.sleep(random_count)

    @classmethod
    def join_path(cls, token, source: str, value: str) -> str:
        return token.join([source, value])

    @classmethod
    def separate_right(cls, value, delimiter) -> str:
        if value is not None:
            x = 0
            x = value.find(delimiter)
            if x != -1:
                x = x + len(delimiter) - 1
                return value[x + 1:len(value)]

    @classmethod
    def separate_left(cls, value, delimiter):
        if value is not None:
            x = 0
            x = value.find(delimiter)
            if x < 1:
                return value
            else:
                return value[0: x]

    @classmethod
    def calc_page(cls, products_count, view_size: int) -> int:
        """
        페이지를 계산해주는 함수
        :arg
            :param products_count: 상품 전체 수
            :param view_size: View Size
        :return:
             self._view_size: 한화면에 로딩하는 상품 개수(20, 40, 60, 80)
        """
        try:
            _page_count = products_count/view_size
        except ZeroDivisionError:
            _page_count = 0

        return math.ceil(_page_count)

    @classmethod
    def export_excel(cls, curs):
        workbook = Workbook()
        file_name = "product_list.xlsx"
        workbook_a = None
        idx = 0
        for data in curs:
            cname: str = data.get('cname')
            cname = cname.replace('/', "")
            if workbook_a is None:
                workbook_a = workbook.create_sheet(index=0, title=cname)
                workbook_a['A1'] = 'title'
                workbook_a['B1'] = 'category'
                workbook_a['C1'] = 'price'
            if workbook_a.title != cname:
                if cname == '바인더1':
                    break
                try:
                    workbook_a = workbook.get_sheet_by_name(cname)
                except KeyError:
                    workbook_a = workbook.create_sheet(index=0, title=cname)

                workbook_a['A1'] = 'title'
                workbook_a['B1'] = 'category'
                workbook_a['C1'] = 'price'
                logging.info('workbook_a - sheet name : ' + workbook_a.title)

            idx += 1
            title = data.get('title')
            price = data.get('price')
            csv_data = (title, cname, price)

            workbook_a.append(csv_data)

            if idx == 1000000:
                logging.info("idx >> {0} ".format(idx))
                idx = 0

        logging.info("csv export start")
        sheets = workbook.get_sheet_names()
        print(sheets)

        workbook.save(file_name)
        logging.info("csv export end")
